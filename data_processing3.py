# -*- coding: utf-8 -*-
"""
@author: Yuri Efremov
import of multiple .dat files and data processing

"""
import numpy as np
import pickle
import pandas as pd
import seaborn as sns  # use boxplot, stripplot, swarmplot, violinplot, catplot
import matplotlib.pyplot as plt
import itertools
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

from Pars_class import Pars_gen
from file_import_qt5simple import file_import_dialog_qt5
from make_Results import make_Results
from utils_ViscoIndent import load_AFM_data_pickle, load_AFM_data_pickle_short


def load_AFM_data_pickle_short2(filename):
    # load
    with open(filename, 'rb') as f:
        data2 = pickle.load(f)
    Pars_dict = data2['Pars']
    if isinstance(Pars_dict, dict):
        Pars = Pars_gen()
        Pars.dict2class(Pars_dict)
    else:
        Pars = Pars_dict
    Data = data2['Data']
    Results = data2['Results']
    return Pars, Data, Results


ngroups = int(input("Select number of groups: "))

ResultsAll = make_Results(0)
ResultsAll['reginds'] = []
counter = 0
group_sizes = []
group_names = []
counterG = 0
for ngroup in range(ngroups):
    group_names.append(input("Name of group %d: " % (ngroup+1)))
for ngroup in range(ngroups):
    filenames = file_import_dialog_qt5(import_opt='multi',
                file_types='*.dat',
                window_title = 'select files for group %d %s' %(ngroup+1, group_names[ngroup]))
    for fname in filenames:
        print(fname)
        Pars, Data, Results = load_AFM_data_pickle_short2(fname)
        if hasattr(Pars.ROIPars, 'reg_nums_all'):
            # reginds = Pars.ROIPars.reg_nums_all[[Pars.ROIPars.reg_nums_all != 0]]
            reginds = Pars.ROIPars.reg_nums_all[Pars.ROIPars.reg_nums_all != 0]
        else:
            reginds = 1
        reginds = reginds + counter
        Results['reginds'] = reginds
        counter = np.max(reginds)
        # ResultsAll = ResultsAll.append(Results, ignore_index=True)
        ResultsAll = pd.concat([ResultsAll, Results], axis=0, join='outer')
        counter2 = int(np.max(ResultsAll['reginds']))
    group_sizes.append(counter2-counterG)
    counterG = counter2

ResultsMean = pd.DataFrame(columns=['Name', 'ROI#', 'mean', 'SD'], index=range(counter2))
col_names = ['Name', 'ROI#', 'mean', 'SD', 'Up_mean', 'Up_SD', 'Down_mean', 'Down_SD', 'Height', 'Number']
ResultsMean = pd.DataFrame(columns=col_names, index=range(counter2))
ResultsMean = ResultsMean.astype({'Name': str,
                                  'ROI#': object,
                                  'mean': np.float32,
                                  'SD': np.float32,
                                  'Up_mean': np.float32,
                                  'Up_SD': np.float32,
                                  'Down_mean': np.float32,
                                  'Down_SD': np.float32,
                                  'Height': np.float32})

keyVs = ['EHertzBEC', 'E0BEC', 'EinfBEC', 'alpha_tauBEC']
# keyVs = ['EHertzBEC', 'Height']
keyVs = ['Eloss']
badinds=[]
for keyV in keyVs:
    for ii in range(1, counter2+1):
        indsc0 = ResultsAll.loc[ResultsAll['reginds'] == ii]
        indsc = indsc0.loc[indsc0['AdjRsq'] > 0.8]  # select AdjRsq level
        if len(indsc)>0:
            indscUp = indsc.loc[indsc['Height'] > np.mean(indsc['Height'])]
            indscDown = indsc.loc[indsc['Height'] < np.mean(indsc['Height'])]
            indsc2 = indsc.loc[indsc['Height'] >= np.percentile(indsc['Height'], 90)]
            height = np.mean(indsc['Height'])
            ResultsMean['Name'].iloc[ii-1] = indsc['Name'].iloc[0]
            ResultsMean['ROI#'].iloc[ii-1] = ii
            ResultsMean['mean'].iloc[ii-1] = np.mean(indsc[keyV])
            ResultsMean['SD'].iloc[ii-1] = np.std(indsc[keyV])
            ResultsMean['Up_mean'].iloc[ii-1] = np.mean(indscUp[keyV])
            ResultsMean['Up_SD'].iloc[ii-1] = np.std(indscUp[keyV])
            ResultsMean['Down_mean'].iloc[ii-1] = np.mean(indscDown[keyV])
            ResultsMean['Down_SD'].iloc[ii-1] = np.std(indscDown[keyV])
            ResultsMean['Height'].iloc[ii-1] = height
            ResultsMean['Number'].iloc[ii-1] = len(indsc[keyV])
        else:
            badinds.append(ii)
    
    data_len = len(ResultsMean.index)
    
    group_list_named = list(itertools.chain.from_iterable(itertools.repeat(x, y) for x, y in zip(group_names, group_sizes)))
    ResultsMean['group_name'] = group_list_named
    
    
    plt.figure()
    ax = sns.swarmplot(x="group_name", y="Up_mean", data=ResultsMean)
    ax = sns.boxplot(x="group_name", y="Up_mean", data=ResultsMean)
    ax.set_title(keyV)
    ax.set_title('')
    ax.set_xlabel('')
    ax.set_ylabel('Height, nm')
    ax.set_ylabel('Young''s modulus, Pa')
    plt.show()

# ResultsMean[["group_name","Up_mean"]].groupby("group_name").mean()
# ResultsMean[["group_name","Up_mean"]].groupby("group_name").describe()
ResultsMean[["group_name", "Up_mean"]].groupby("group_name").agg(['median', 'mean', 'std'])
Res2 = ResultsMean[["group_name", "mean"]].groupby("group_name").agg(['median', 'mean', 'std'])

#  filename = 'D:/MEGAsync/My materials/python/Ting_code/examples/Bruker_forcevolume_cells.dat'
filenamexls = filenames[1]+'.xlsx'
# ResultsMean.to_excel(filenamexls) # deprecated
wb = Workbook()
ws = wb.active

for r in dataframe_to_rows(ResultsMean, index=True, header=True):
    ws.append(r)

for cell in ws['A'] + ws[1]:
    cell.style = 'Pandas'

wb.save(filenamexls)
