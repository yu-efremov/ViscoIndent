# -*- coding: utf-8 -*-
"""
@author: Yuri Efremov
import of multiple .dat files and data processing

"""
import os
import numpy as np
import pickle
import pandas as pd
import seaborn as sns  # use boxplot, stripplot, swarmplot, violinplot, catplot
import matplotlib.pyplot as plt
import itertools
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

from Pars_class import Pars_gen
from file_import_qt5simple import file_import_dialog_qt5
from make_Results import make_Results
from utils_ViscoIndent import load_AFM_data_pickle, load_AFM_data_pickle_short


def load_AFM_data_pickle_short2(filename):
    # load
    with open(filename, 'rb') as f:
        data2 = pickle.load(f)
    Pars_dict = data2['Pars']
    if isinstance(Pars_dict, dict):
        Pars = Pars_gen()
        Pars.dict2class(Pars_dict)
    else:
        Pars = Pars_dict
    Data = data2['Data']
    Results = data2['Results']
    return Pars, Data, Results


ngroups = int(input("Select number of groups: "))

ResultsAll = make_Results(0)
ResultsAll['reginds'] = []
counter = 0
group_sizes = []
group_names = []
counterG = 0
for ngroup in range(ngroups):
    group_names.append(input("Name of group %d: " % (ngroup+1)))
for ngroup in range(ngroups):
    filenames = file_import_dialog_qt5(import_opt='multi',
                file_types='*.dat',
                window_title = 'select files for group %d %s' %(ngroup+1, group_names[ngroup]))
    for fname in filenames:
        print(fname)
        Pars, Data, Results = load_AFM_data_pickle_short2(fname)
        if hasattr(Pars.ROIPars, 'reg_nums_all'):
            # reginds = Pars.ROIPars.reg_nums_all[[Pars.ROIPars.reg_nums_all != 0]]
            reginds = Pars.ROIPars.reg_nums_all[Pars.ROIPars.reg_nums_all != 0]
        else:
            reginds = 1
        reginds = reginds + counter
        Results['reginds'] = reginds
        counter = np.max(reginds)
        # ResultsAll = ResultsAll.append(Results, ignore_index=True)
        ResultsAll = pd.concat([ResultsAll, Results], axis=0, join='outer')
        counter2 = int(np.max(ResultsAll['reginds']))  # total nuber of ROIs
    group_sizes.append(counter2-counterG)
    counterG = counter2
    
def get_mean_results(ResultsAll, keyV, group_names, height_analysis):
    counter2 = int(np.max(ResultsAll['reginds']))  # total nuber of ROIs
    badinds=[]
    if height_analysis==1:
        col_names = ['Name', 'ROI#', 'mean', 'SD', 'Up_mean', 'Up_SD', 'Down_mean', 'Down_SD', 'Height', 'Number']
        ResultsMean = pd.DataFrame(columns=col_names, index=range(counter2))
        ResultsMean = ResultsMean.astype({'Name': str,
                                          'ROI#': object,
                                          'mean': np.float64,
                                          'SD': np.float64,
                                          'Up_mean': np.float64,
                                          'Up_SD': np.float64,
                                          'Down_mean': np.float64,
                                          'Down_SD': np.float64,
                                          'Height': np.float64})
    elif height_analysis==0:
        col_names = ['Name', 'ROI#', 'mean', 'SD', 'Height', 'Number']
        ResultsMean = pd.DataFrame(columns=col_names, index=range(counter2))
        ResultsMean = ResultsMean.astype({'Name': str,
                                          'ROI#': object,
                                          'mean': np.float64,
                                          'SD': np.float64,
                                          'Height': np.float64})
    if keyV == 'Height':
        for ii in range(1, counter2+1):
            indsc = ResultsAll.loc[ResultsAll['reginds'] == ii]
            if len(indsc)>0:
                ResultsMean.loc[ii-1, 'Name'] = indsc['Name'].iloc[0]
                ResultsMean.loc[ii-1, 'ROI#'] = ii
                indsc2 = indsc.loc[indsc['Height'] >= np.percentile(indsc['Height'], 80)]
                height = np.mean(indsc2['Height'])
                heighSD = np.std(indsc2['Height'])
                if len(indsc2)==0:
                    height = np.mean(indsc['Height'])
                    heighSD = np.std(indsc['Height'])
                ResultsMean.loc[ii-1, 'mean'] = height
                ResultsMean.loc[ii-1, 'SD'] = heighSD
            else:
                badinds.append(ii)
    else:  # all other keyV
        for ii in range(1, counter2+1):
            indsc = ResultsAll.loc[ResultsAll['reginds'] == ii]
            if len(indsc)>0:
                ResultsMean.loc[ii-1, 'Name'] = indsc['Name'].iloc[0]
                ResultsMean.loc[ii-1, 'ROI#'] = ii
                ResultsMean.loc[ii-1, 'mean'] = np.mean(indsc[keyV])
                ResultsMean.loc[ii-1, 'SD'] = np.std(indsc[keyV])
                if height_analysis==1:
                    indscUp = indsc.loc[indsc['Height'] > np.mean(indsc['Height'])]
                    indscDown = indsc.loc[indsc['Height'] < np.mean(indsc['Height'])]
                    ResultsMean.loc[ii-1, 'Up_mean'] = np.mean(indscUp[keyV])
                    ResultsMean.loc[ii-1, 'Up_SD'] = np.std(indscUp[keyV])
                    ResultsMean.loc[ii-1, 'Down_mean'] = np.mean(indscDown[keyV])
                    ResultsMean.loc[ii-1, 'Down_SD'] = np.std(indscDown[keyV])
                    ResultsMean.loc[ii-1, 'Number'] = len(indsc[keyV])
            else:
                badinds.append(ii)
                
    # data_len = len(ResultsMean.index)
    group_list_named = list(itertools.chain.from_iterable(itertools.repeat(x, y) for x, y in zip(group_names, group_sizes)))
    ResultsMean['group_name'] = group_list_named
    ResultsTotal = pd.DataFrame(columns=[group_names], index=range(max(group_sizes)))
    ResultsUp = pd.DataFrame(columns=[group_names], index=range(max(group_sizes)))
    ResultsDown = pd.DataFrame(columns=[group_names], index=range(max(group_sizes)))
    
    for group_name in group_names:
        Total_mean = ResultsMean.loc[ResultsMean["group_name"] == group_name, "mean"]
        Total_mean.index = Total_mean.index - Total_mean.index.min()
        ResultsTotal[group_name] = Total_mean
        if height_analysis==1:
            Up_mean = ResultsMean.loc[ResultsMean["group_name"] == group_name, "Up_mean"]
            Up_mean.index = Up_mean.index - Up_mean.index.min()
            ResultsUp[group_name] = Up_mean
            Down_mean = ResultsMean.loc[ResultsMean["group_name"] == group_name, "Down_mean"]
            Down_mean.index = Down_mean.index - Down_mean.index.min()
            ResultsDown[group_name] = Down_mean       
        # nan_idx = ResultsYM[group_name].isna() & ResultsH[group_name].isna()
        # idx_list = nan_idx[group_name].values.tolist()
        # nan_idx = np.where(idx_list)[0]
        # ResultsYM[group_name] = ResultsYM[group_name].drop(index=nan_idx).reset_index(drop=True)
        # ResultsH[group_name] =  ResultsH[group_name].drop(index=nan_idx).reset_index(drop=True)

    return ResultsMean, ResultsTotal, ResultsUp, ResultsDown, badinds

ResultsMeanYM, _, ResultsYM, _, badinds = get_mean_results(ResultsAll, 'EHertzBEC', group_names, height_analysis=1)
ResultsMeanH, ResultsH, _, _, badinds = get_mean_results(ResultsAll, 'Height', group_names, height_analysis=1)
ResultsMeanE0BEC, _, ResultsE0BEC, _, badinds = get_mean_results(ResultsAll, 'E0BEC', group_names, height_analysis=1)
ResultsMeanEinfBEC, _, ResultsEinfBEC, _, badinds = get_mean_results(ResultsAll, 'EinfBEC', group_names, height_analysis=1)
ResultsMeanalpha_tauBEC, _, Resultsalpha_tauBEC, _, badinds = get_mean_results(ResultsAll, 'alpha_tauBEC', group_names, height_analysis=1)
ResultsH = ResultsH/1000; ResultsHnp = ResultsH.to_numpy() # to um, to Prizm
ResultsYM = ResultsYM/1000; ResultsYMnp = ResultsYM.to_numpy() # to kPa
ResultsE0BEC = ResultsE0BEC/1000; ResultsE0BECnp = ResultsE0BEC.to_numpy() # to kPa
ResultsEinfBECnp = ResultsEinfBEC.to_numpy()
Resultsalpha_tauBECnp = Resultsalpha_tauBEC.to_numpy()

fig, axes = plt.subplots(1, 2, figsize=(12, 5), sharex=True)
# --- First subplot: Height ---
sns.swarmplot(x="group_name", y="mean", data=ResultsMeanH, ax=axes[0])
sns.boxplot(x="group_name", y="mean", data=ResultsMeanH, ax=axes[0])
axes[0].set_title('Height')
axes[0].set_xlabel('')
axes[0].set_ylabel('Height, um')
axes[0].set_ylim(bottom=0) 
# --- Second subplot: Young's modulus ---
sns.swarmplot(x="group_name", y="Up_mean", data=ResultsMeanYM, ax=axes[1])
sns.boxplot(x="group_name", y="Up_mean", data=ResultsMeanYM, ax=axes[1])
axes[1].set_title("Young's modulus")
axes[1].set_xlabel('')
axes[1].set_ylabel("Young's modulus, Pa")
axes[1].set_ylim(bottom=0) 
plt.tight_layout()
plt.show()

viscoelastic_plots = 1
if viscoelastic_plots == 1:
    fig2, axes2 = plt.subplots(1, 3, figsize=(12, 5), sharex=True)
    # --- First subplot: E0 ---
    sns.swarmplot(x="group_name", y="Up_mean", data=ResultsMeanE0BEC, ax=axes2[0])
    sns.boxplot(x="group_name", y="Up_mean", data=ResultsMeanE0BEC, ax=axes2[0])
    axes2[0].set_title("E1")
    axes2[0].set_xlabel('')
    axes2[0].set_ylabel("E1, Pa")
    axes2[0].set_ylim(bottom=0) 
    # --- Second subplot: alpha_tau ---
    sns.swarmplot(x="group_name", y="Up_mean", data=ResultsMeanalpha_tauBEC, ax=axes2[1])
    sns.boxplot(x="group_name", y="Up_mean", data=ResultsMeanalpha_tauBEC, ax=axes2[1])
    axes2[1].set_title("alpha")
    axes2[1].set_xlabel('')
    axes2[1].set_ylabel("alpha")
    axes2[1].set_ylim(bottom=0) 
    # --- Third subplot: Einf ---
    sns.swarmplot(x="group_name", y="Up_mean", data=ResultsMeanEinfBEC, ax=axes2[2])
    sns.boxplot(x="group_name", y="Up_mean", data=ResultsMeanEinfBEC, ax=axes2[2])
    axes2[2].set_title("Viscosity")
    axes2[2].set_xlabel('')
    axes2[2].set_ylabel("Viscosity")
    axes2[2].set_ylim(bottom=0) 
    plt.tight_layout()
    plt.show()

# ResultsMean[["group_name","Up_mean"]].groupby("group_name").mean()
# ResultsMean[["group_name","Up_mean"]].groupby("group_name").describe()
# ResultsMeanYM[["group_name", "Up_mean"]].groupby("group_name").agg(['median', 'mean', 'std'])
ResYM_groupmean = ResultsMeanYM[["group_name", "mean"]].groupby("group_name").agg(['median', 'mean', 'std'])
ResH_groupmean = ResultsMeanH[["group_name", "mean"]].groupby("group_name").agg(['median', 'mean', 'std'])

#  filename = 'D:/MEGAsync/My materials/python/Ting_code/examples/Bruker_forcevolume_cells.dat'
dirname = os.path.dirname(filenames[0])
foldername = os.path.basename(dirname)
filenamexls = dirname + "/" + foldername + '.xlsx'
# ResultsMean.to_excel(filenamexls) # deprecated
wb = Workbook()
ws = wb.active
ResultsMeanYM['Name'] = ResultsMeanYM['Name'].astype('str') 
for r in dataframe_to_rows(ResultsMeanYM, index=True, header=True):
    ws.append(r)

for cell in ws['A'] + ws[1]:
    cell.style = 'Pandas'

# wb.save(filenamexls)
